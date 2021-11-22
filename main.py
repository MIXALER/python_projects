import os
import xlrd
import openpyxl
import re

# 第二张表和第三张表都从第二行开始读数据直到倒数第二行
input_excel = r"C:\Users\yuanh\Desktop\input"

output_excel = r"C:\Users\yuanh\Desktop\output\生成结果.xlsx"

title = ['科室', '核定床', '占床', '空床', '入', '出']

fuke = ["妇产科一病区（产科）", "妇产科二病区（妇科）", "妇产科三病区", "产房"]

wuguan = ["眼•耳鼻喉科二病区", "眼•耳鼻喉科", "耳鼻咽喉科门诊", '眼科门诊']

erke = ['儿科', '儿科icu', '儿保']

kouqiang = ['口腔科']

ganranke = ['感染科']

quanke = ['全科医学科']

zhongzheng = ['重症医学科']

kangfuke = ['康复科']

xueyeke = ['血液透析中心（透析台次）']

pifuke = []

jizhengke = []

zhongyike = []

templates_setting = [fuke, wuguan, erke, kouqiang]


def read_files(file_path):
    all_files = list()

    all_file_names = os.listdir(file_path)

    for i_name in all_file_names:
        all_files.append(os.path.join(file_path, i_name))

    return all_files


def read_excel_one(excel):
    excel_data = xlrd.open_workbook(excel)
    table = excel_data.sheet_by_index(1)
    nrows, ncols = table.nrows, table.ncols
    ret = {}

    for i in range(1, nrows):
        ret[str(table.cell(i, 0).value)] = int(table.cell(i, 1).value)
    return ret


def read_excel_two(excel):
    excel_data = xlrd.open_workbook(excel)
    table = excel_data.sheet_by_index(2)
    nrows, ncols = table.nrows, table.ncols
    ret = {}

    for i in range(2, nrows):
        ret[str(table.cell(i, 0).value)] = [int(table.cell(i, 1).value), int(table.cell(i, 2).value),
                                            int(table.cell(i, 3).value), int(table.cell(i, 4).value),
                                            int(table.cell(i, 5).value)]
    return ret


def write_results_to_excel(my_template, ret1, ret2):
    work_book = openpyxl.Workbook()

    for index, temp in enumerate(my_template):
        work_sheet = work_book.create_sheet(index=index)
        for i, item in enumerate(title, 1):
            work_sheet.cell(1, i, item)

        count = 2
        for i in temp:
            if i not in ret2:
                continue
            for j, num in enumerate(ret2[i], 1):
                work_sheet.cell(count, j, num)
                count = count + 1
    work_book.save(output_excel)


if __name__ == '__main__':
    print('读取 excel 文件')
    all_excel_files = read_files(input_excel)
    print('excel文件数{}'.format(len(all_excel_files)))

    for excel_file in all_excel_files:
        ret1 = read_excel_one(excel_file)
        ret2 = read_excel_two(excel_file)
        print('写为excel表')
        write_results_to_excel(templates_setting, ret1, ret2)

    print('完毕')
