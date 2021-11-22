import os
import pandas as pd
import xlrd
import openpyxl
import re
# import docx
# from tools.validate import check_email

CSV_FILES_PATH = r'D:\teacher_wen\原始文件'
EXCEL_FILES_SAVING_PATH = r'D:\teacher_wen\原始文件\excels'

EXCEL_FILES_PATH = r'D:\teacher_wen\待发送\isaic\中文'
EXCEL_SAVE_FILE = r'D:\teacher_wen\待发送\isaic\isaic_Chinese_20210729.xlsx'

SEARCH_STR = '电子邮件:'

already_exist_emails = dict()
needed_email_info = list()


def read_files(file_path):
    all_files = list()

    all_file_names = os.listdir(file_path)

    for i_name in all_file_names:
        all_files.append(os.path.join(file_path, i_name))

    return all_files


def read_word(filepath):
    # 输出一个列表，元素为 word 里的一个段落
    file = docx.Document(filepath)
    print("段落数:" + str(len(file.paragraphs)))  # 段落数为13，每个回车隔离一段
    excel_list = list()
    # 输出每一段的内容
    for para in file.paragraphs:
        print(para.text)
        excel_list.append(para.text)
    return excel_list


def csv_to_excel(csv_file, excel_files_saving_path, index_excel):
    readed_csv = pd.read_csv(csv_file, encoding='UTF-8', error_bad_lines=False)
    readed_csv.to_excel(excel_files_saving_path + '{}.xlsx'.format(index_excel))


def seach_and_save(excel_file, search_str='email', index_in_tabels=0):
    ce = 0
    cn = 0
    ct = 0
    ci = 0

    excel_data = xlrd.open_workbook(excel_file)
    table = excel_data.sheet_by_index(index_in_tabels)
    nrows, ncols = table.nrows, table.ncols

    first_row = table.row_values(0, 0, ncols)
    first_row_search_counter = 0
    for i_cell in first_row:

        if i_cell == "作者":
            cn = first_row_search_counter
            first_row_search_counter += 1
            continue
        if i_cell == "摘要":
            ci = first_row_search_counter
            first_row_search_counter += 1
            continue
        if i_cell == "通讯地址":
            ce = first_row_search_counter
            first_row_search_counter += 1
            continue
        if i_cell == "标题":
            ct = first_row_search_counter
            first_row_search_counter += 1
            continue

        first_row_search_counter += 1

    for i_row in range(nrows):
        if re.search(search_str, str(table.cell(i_row, ce).value)):
            for_check_email = table.cell(i_row, ce).value.split(search_str)[1][1:]
            if check_email(for_check_email):
                if table.cell(i_row, ce).value.split(search_str)[1] not in already_exist_emails:
                    already_exist_emails[table.cell(i_row, ce).value.split(search_str)[1]] = None
                    needed_email_info.append([table.cell(i_row, ce).value.split(search_str)[1],
                                              table.cell(i_row, cn).value.split('.,')[0],
                                              table.cell(i_row, ct).value,
                                              table.cell(i_row, ci).value])


def merge_excels(excels):
    info = []
    for excel_file in excels:
        excel_data = xlrd.open_workbook(excel_file)
        table = excel_data.sheet_by_index(0)
        nrows, ncols = table.nrows, table.ncols
        for i_row in range(nrows):
            if table.cell(i_row, 0).value not in already_exist_emails:
                already_exist_emails[table.cell(i_row, 0).value] = None
                info.append([table.cell(i_row, 0).value.strip(),
                             table.cell(i_row, 1).value,
                             table.cell(i_row, 2).value,
                             table.cell(i_row, 3).value])

    return info


def write_results_to_excel(result_list, save_file, ce=1, cn=2, ct=3, ci=4):
    work_book = openpyxl.Workbook()
    work_sheet = work_book.create_sheet(title='result', index=0)
    for index in range(len(result_list)):
        work_sheet.cell(index + 1, ce, result_list[index][0].strip())
        work_sheet.cell(index + 1, cn, result_list[index][1])
        work_sheet.cell(index + 1, ct, result_list[index][2])
        work_sheet.cell(index + 1, ci, result_list[index][3])
    work_book.save(save_file)


if __name__ == '__main__':
    # all_csv_files = read_files(CSV_FILES_PATH)
    #
    # print('csv文件总数{}'.format(len(all_csv_files)))
    #
    # index_transfer = 1
    #
    # for i_csv_file in all_csv_files:
    #     print('当前转换{}'.format(i_csv_file))
    #
    #     csv_to_excel(i_csv_file, EXCEL_FILES_SAVING_PATH, index_transfer)
    #     index_transfer += 1
    #
    # print('csv文件内容转换完毕')

    all_excel_files = read_files(EXCEL_FILES_PATH)

    print('excel文件数{}'.format(len(all_excel_files)))

    search_counter = 0
    # for i_excel_file in all_excel_files:
    #     already_found = len(needed_email_info)
    #
    #     seach_and_save(i_excel_file, search_str=SEARCH_STR, index_in_tabels=0)
    #     search_counter += 1
    #
    #     print('当前搜索第{}个excel表, 表：{}'.format(search_counter, i_excel_file))
    #     print('当前找到{}条邮件信息，本excel表{}条邮件信息'.format(len(needed_email_info),
    #                                               len(needed_email_info) - already_found))

    # 合并 excel 文件
    needed_email_info = merge_excels(all_excel_files)

    print("邮箱总条数{}".format(len(already_exist_emails)))
    print("邮件信息总条数{}".format(len(needed_email_info)))
    print('写为excel表')
    write_results_to_excel(needed_email_info, save_file=EXCEL_SAVE_FILE)

    print('完毕')
